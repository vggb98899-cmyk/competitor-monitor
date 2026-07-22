"""
工具模块：存放Excel、日志、价格历史等通用功能
"""
import json
import logging
from pathlib import Path
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─── 日志设置（同时输出到控制台+文件） ───
LOG_DIR = Path(__file__).parent / "logs"
LOG_DIR.mkdir(exist_ok=True)
log_file = LOG_DIR / f"{date.today().isoformat()}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),  # 控制台
        logging.FileHandler(log_file, encoding="utf-8"),  # 文件
    ],
)
logger = logging.getLogger(__name__)
logger.info(f"📝 日志文件: {log_file}")


# ─── 价格历史读写 ───

def load_price_history(filepath: Path) -> dict:
    """
    读取价格历史文件，返回 { "日期": { "商品标题": "价格", ... }, ... }
    如果文件不存在，返回空字典
    """
    if not filepath.exists():
        logger.info("价格历史文件不存在，首次运行")
        return {}
    with open(filepath, "r", encoding="utf-8") as f:
        history = json.load(f)
    logger.info(f"已读取价格历史，共 {len(history)} 天记录")
    return history


def save_price_history(filepath: Path, history: dict) -> None:
    """保存价格历史到文件"""
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)
    logger.info(f"价格历史已保存 → {filepath}")


def attach_price_changes(
    products: list[dict],
    history: dict,
    today_str: str,
    yesterday_str: str,
) -> list[dict]:
    """
    给每个商品加上"相比昨日涨跌额"字段

    Args:
        products: 商品列表（每项含 title / price）
        history: 历史数据 { "2026-07-08": { "商品标题": "价格", ... }, ... }
        today_str: 今天的日期字符串 "2026-07-09"
        yesterday_str: 昨天的日期字符串 "2026-07-08"

    Returns:
        新增了 change / change_label 字段的商品列表
    """
    yesterday_prices = history.get(yesterday_str, {})

    for product in products:
        title = product["title"]
        current_price = product["price"]

        # 如果没有历史价格
        if not yesterday_prices or title not in yesterday_prices:
            product["change"] = ""
            product["change_label"] = "—"
            continue

        # 计算涨跌
        try:
            old_price = float(yesterday_prices[title])
            new_price = float(current_price)
            diff = round(new_price - old_price, 2)

            if diff > 0:
                product["change_label"] = f"+${diff:.2f} ↑"
            elif diff < 0:
                product["change_label"] = f"-${abs(diff):.2f} ↓"
            else:
                product["change_label"] = "$0.00 ="
            product["change"] = diff
        except (ValueError, TypeError):
            product["change"] = ""
            product["change_label"] = "—"

    changed_count = sum(1 for p in products if p.get("change_label") not in ("—", ""))
    logger.info(f"价格对比完成，{changed_count} 个商品有历史可对比")
    return products


def update_history(
    history: dict,
    products: list[dict],
    today_str: str,
) -> dict:
    """
    将今天的商品价格写入历史记录

    Args:
        history: 现有历史数据
        products: 今天的商品列表
        today_str: 今天日期

    Returns:
        更新后的历史数据
    """
    today_record = {}
    for p in products:
        today_record[p["title"]] = p["price"]

    history[today_str] = today_record
    logger.info(f"已记录今天 ({today_str}) 共 {len(today_record)} 个商品价格")
    return history


# ─── Excel 输出 ───

def save_to_excel(products: list[dict], file_path: Path, category_summary: list[dict] | None = None) -> str:
    """
    将商品列表保存为 Excel 文件

    Args:
        products: 商品字典列表，每项含 title / price / sales / category / change_label
        file_path: 输出文件路径

    Returns:
        保存成功的文件路径字符串
    """
    # 确保输出目录存在
    file_path.parent.mkdir(parents=True, exist_ok=True)

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品数据"

    # ─── 表头 ───
    headers = ["商品标题", "品类", "价格(USD)", "月销量", "相比昨日涨跌额"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ─── 数据行 ───
    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.get("title", ""))
        ws.cell(row=row, column=2, value=product.get("category", ""))
        ws.cell(row=row, column=3, value=product.get("price", ""))
        ws.cell(row=row, column=4, value=product.get("sales", ""))
        ws.cell(row=row, column=5, value=product.get("change_label", ""))

    # ─── 调整列宽 ───
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20

    # ─── 品类排行 Sheet（如果有数据） ───
    if category_summary:
        ws2 = wb.create_sheet(title="品类排行", index=0)  # 插到最前面

        # 表头
        headers2 = ["排名", "品类", "商品数量"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        for row, item in enumerate(category_summary, 2):
            ws2.cell(row=row, column=1, value=row - 1)  # 排名
            ws2.cell(row=row, column=2, value=item["category"])
            ws2.cell(row=row, column=3, value=item["count"])

        ws2.column_dimensions["A"].width = 8
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 12

    # 保存
    wb.save(file_path)
    logger.info(f"Excel 已保存 → {file_path}")
    return str(file_path)


def save_probe_report(results: list[dict], file_path: Path) -> str:
    """
    输出《38家店铺数据可信度探测报告》

    Args:
        results: probe_store() 返回的结果列表
        file_path: 输出Excel路径

    Returns:
        保存成功的文件路径
    """
    file_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "探测报告"

    # 表头
    headers = [
        "店铺名称", "店铺URL", "主力品类", "主力店",
        "接口是否可访问", "是否返回月销量", "月销量字段名",
        "页面是否显示销量文本", "销量文本内容",
        "是否有评分", "评分相关内容",
        "数据等级预判", "备注",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # 数据行
    for row, r in enumerate(results, 2):
        api = r.get("api", {})
        page = r.get("page", {})

        ws.cell(row=row, column=1, value=r.get("name", ""))
        ws.cell(row=row, column=2, value=r.get("url", ""))
        ws.cell(row=row, column=3, value=r.get("category", ""))
        ws.cell(row=row, column=4, value="⭐是" if r.get("is_key") else "")

        # 接口相关
        ws.cell(row=row, column=5, value="✅" if api.get("api_accessible") else "❌")
        ws.cell(row=row, column=6, value="✅" if api.get("has_sales_field") else "❌")
        ws.cell(row=row, column=7, value=api.get("sales_field_name", ""))

        # 页面相关
        ws.cell(row=row, column=8, value="✅" if page.get("has_sales_text") else "❌")
        ws.cell(row=row, column=9, value=page.get("sales_text", ""))
        ws.cell(row=row, column=10, value="✅" if page.get("has_rating") else "❌")
        ws.cell(row=row, column=11, value=page.get("rating_text", ""))

        # 等级
        level = r.get("data_level", "C")
        level_label = {"A": "A类（有销量）", "B": "B类（有评分）", "C": "C类（基础字段）"}
        ws.cell(row=row, column=12, value=level_label.get(level, level))

        # 备注（收集错误信息）
        notes = []
        if api.get("error"):
            notes.append(f"接口: {api['error']}")
        if page.get("error"):
            notes.append(f"页面: {page['error']}")
        ws.cell(row=row, column=13, value="; ".join(notes) if notes else "")

    # 调整列宽
    col_widths = [16, 35, 12, 8, 14, 14, 14, 18, 30, 12, 30, 16, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(file_path)
    logger.info(f"📊 探测报告已保存 → {file_path}")
    return str(file_path)


# ─── 评分历史 ───

def load_rating_history(filepath: Path) -> dict:
    """
    读取评分历史文件
    结构: { "店铺名": { "商品标题": { "日期": "评分", ... } }, ... }
    """
    if not filepath.exists():
        logger.info("评分历史文件不存在，首次运行")
        return {}
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    logger.info(f"已读取评分历史，共 {len(data)} 家店铺")
    return data


def save_rating_history(filepath: Path, history: dict) -> None:
    """保存评分历史"""
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def get_rating_trend(rating_history: dict, store: str, title: str, days: int = 7) -> str:
    """
    获取某个商品最近N天的评分趋势

    Returns:
        如 "4.5 → 4.5 → 4.3"（按日期从旧到新）
    """
    store_history = rating_history.get(store, {})
    product_history = store_history.get(title, {})

    if not product_history:
        return ""

    # 按日期排序，取最近N天
    sorted_dates = sorted(product_history.keys())[-days:]
    values = [product_history[d] for d in sorted_dates]

    # 去重尾部连续相同值，只保留变化
    # 但如果只有1个值也显示
    if len(values) <= 1:
        return values[0] if values else ""

    # 显示所有值
    return " → ".join(values)


# ─── 新品发现跟踪 ───

def load_discovery_history(filepath: Path) -> dict:
    """
    读取新品发现历史
    结构: { "店铺名": { "商品标题": "首次发现日期", ... }, ... }
    """
    if not filepath.exists():
        logger.info("新品发现文件不存在，首次运行")
        return {}
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    logger.info(f"已读取新品发现记录，共 {len(data)} 家店铺")
    return data


def save_discovery_history(filepath: Path, history: dict) -> None:
    """保存新品发现历史"""
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def update_discovery(
    discovery: dict,
    products: list[dict],
    store_name: str,
    today_str: str,
) -> dict:
    """
    更新新品发现时间

    逻辑：
      - 如果商品已存在 → 保留首次发现时间
      - 如果商品是新的 → 记录今天为首次发现时间

    Returns:
        更新后的 discovery 字典
    """
    store_discovery = discovery.get(store_name, {})

    for p in products:
        title = p.get("title", "")
        if not title:
            continue
        if title not in store_discovery:
            store_discovery[title] = today_str

    discovery[store_name] = store_discovery
    return discovery


# ─── 竞品分析 Excel 输出 ───

def save_competitor_excel(
    all_products: list[dict],
    file_path: Path,
    rating_history: dict,
    discovery_history: dict,
    price_changes: dict,
) -> str:
    """
    输出竞品分析Excel

    Args:
        all_products: 所有商品列表，
            每项含 title / price / category / status / store / handle / rating
        file_path: 输出路径
        rating_history: 评分历史
        discovery_history: 新品发现历史
        price_changes: 价格变化 { "商品标题": "涨跌额标签" }

    Returns:
        保存路径
    """
    file_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "竞品数据"

    # ─── 表头 ───
    headers = [
        "商品标题", "店铺", "品类", "当前价格",
        "商品状态", "评分", "相比昨日涨跌额",
        "评分7天趋势", "首次发现时间",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # ─── 数据行 ───
    for row, p in enumerate(all_products, 2):
        title = p.get("title", "")
        store = p.get("store", "")
        change_label = price_changes.get(f"{store}::{title}", "—")
        rating_trend = get_rating_trend(rating_history, store, title)
        discovered = discovery_history.get(store, {}).get(title, "")

        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=2, value=store)
        ws.cell(row=row, column=3, value=p.get("category", ""))
        ws.cell(row=row, column=4, value=p.get("price", ""))
        ws.cell(row=row, column=5, value=p.get("status", ""))
        ws.cell(row=row, column=6, value=p.get("rating", ""))
        ws.cell(row=row, column=7, value=change_label)
        ws.cell(row=row, column=8, value=rating_trend)
        ws.cell(row=row, column=9, value=discovered)

    # ─── 列宽 ───
    widths = [50, 14, 14, 12, 10, 8, 18, 30, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(file_path)
    logger.info(f"📊 竞品分析已保存 → {file_path}")
    return str(file_path)
